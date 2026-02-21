"""
SLR Excel to VF Store Importer
==============================
Import systematic literature review data from Excel to Veritas VF Store.

Usage:
    python scripts/import_slr_excel.py

Developed by Lily Xiaolei Li
"""

import json
import os
import sys
from pathlib import Path

# Add backend to path
sys.path.insert(0, str(Path(__file__).parent.parent / "backend"))

# Excel file path
EXCEL_FILE = r"C:\Users\thene\OneDrive - The University Of Newcastle\Desktop\Newly Structured Folder\Paper 1\Data\Paper 1 SLR data and analysis (5 Jan 2026).xlsx"
SHEET_NAME = "NFIA_data"

# Qdrant configuration - use local data directory
QDRANT_PATH = Path(__file__).parent.parent / "data" / "qdrant_slr"
COLLECTION_NAME = "vf_profiles_slr"


import argparse


def read_excel_data(start_index: int | None = None, end_index: int | None = None):
    """Read Excel file using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Installing openpyxl...")
        os.system(f"{sys.executable} -m pip install openpyxl")
        from openpyxl import load_workbook

    print(f"📖 Reading Excel: {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    sheet = wb[SHEET_NAME]

    # Get headers from first row
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value if cell.value else f"col_{len(headers)}")

    print(f"   Found {len(headers)} columns")

    # Read data rows
    papers = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Skip empty rows
            continue

        paper = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers):
                paper[headers[col_idx]] = value

        # Optional index range filter (use dataset 'index' column)
        try:
            paper_index = int(paper.get("index"))
        except (TypeError, ValueError):
            paper_index = None

        if start_index is not None and paper_index is not None and paper_index < start_index:
            continue
        if end_index is not None and paper_index is not None and paper_index > end_index:
            continue

        papers.append(paper)

    wb.close()
    print(f"   Found {len(papers)} papers")
    return papers


def paper_to_vf_profile(paper: dict) -> dict:
    """Convert Excel paper data to VF profile format."""

    # Generate paper_id from index
    paper_id = f"slr_{paper.get('index', 'unknown')}"

    # Build meta chunk
    meta = {
        "index": paper.get("index"),
        "authors": paper.get("AUTHORS", ""),
        "year": paper.get("year"),
        "title": paper.get("title", ""),
        "journal": paper.get("journal", ""),
        "doi": paper.get("doi", ""),
        "volume": paper.get("volume_(Issue)", ""),
        "pages": paper.get("pages", ""),
        "jnl_rank": paper.get("jnl_rank", ""),
        "paper_type": paper.get("paper_type", ""),
        "country": paper.get("country", ""),
        "source": paper.get("source", ""),
    }

    # Build 8 semantic chunks
    chunks = {
        "meta": json.dumps(meta, ensure_ascii=False, default=str),

        "abstract": str(paper.get("abstract", "") or ""),

        "theory": str(paper.get("theory", "") or ""),

        "literature": _combine_fields([
            paper.get("primary_literature"),
            paper.get("other_literature"),
        ]),

        "research_questions": _combine_fields([
            paper.get("rq_hypothesis_1"),
            paper.get("rq_hypothesis_2"),
            paper.get("rq_hypothesis_3"),
        ]),

        "contributions": _combine_fields([
            paper.get("contributions"),
            paper.get("finding_1"),
            paper.get("finding_2"),
            paper.get("finding_3"),
        ]),

        "key_concepts": _combine_fields([
            paper.get("AUTHOR KEYWORDS"),
            paper.get("INDEX KEYWORDS"),
            paper.get("theme_primary"),
            paper.get("sub_theme"),
        ]),

        "cited_for": _combine_fields([
            paper.get("motivation"),
            paper.get("research_gap"),
            paper.get("tensions_debate"),
        ]),
    }

    return {
        "paper_id": paper_id,
        "in_library": True,
        "source_type": "slr_excel",
        "meta": meta,
        "chunks": chunks,
    }


def _combine_fields(fields: list) -> str:
    """Combine multiple fields into one text, filtering None/empty."""
    parts = []
    for f in fields:
        if f and str(f).strip():
            parts.append(str(f).strip())
    return " | ".join(parts)


def setup_qdrant():
    """Initialize Qdrant client and collection."""
    try:
        from qdrant_client import QdrantClient
        from qdrant_client.http.models import Distance, VectorParams
    except ImportError:
        print("Installing qdrant-client...")
        os.system(f"{sys.executable} -m pip install qdrant-client")
        from qdrant_client import QdrantClient
        from qdrant_client.http.models import Distance, VectorParams

    # Connect to Qdrant server
    QDRANT_URL = os.getenv("QDRANT_URL", "http://localhost:6333")
    print(f"🗄️  Connecting to Qdrant at {QDRANT_URL}")
    client = QdrantClient(url=QDRANT_URL)

    # Create collection if not exists
    collections = [c.name for c in client.get_collections().collections]
    if COLLECTION_NAME not in collections:
        client.create_collection(
            collection_name=COLLECTION_NAME,
            vectors_config=VectorParams(size=1024, distance=Distance.COSINE),
        )
        print(f"   Created collection: {COLLECTION_NAME}")
    else:
        print(f"   Collection exists: {COLLECTION_NAME}")

    return client


def get_embedding_model():
    """Load BGE-M3 embedding model."""
    try:
        from sentence_transformers import SentenceTransformer
    except ImportError:
        print("Installing sentence-transformers...")
        os.system(f"{sys.executable} -m pip install sentence-transformers")
        from sentence_transformers import SentenceTransformer

    print("🧠 Loading BGE-M3 embedding model...")
    model = SentenceTransformer("BAAI/bge-m3")
    print("   Model loaded!")
    return model


def import_to_vf_store(papers: list, client, model):
    """Import papers to VF Store."""
    import uuid
    from qdrant_client.models import PointStruct

    VF_NAMESPACE = uuid.UUID("a1b2c3d4-e5f6-7890-abcd-ef1234567890")
    CHUNK_ORDER = ["meta", "abstract", "theory", "literature",
                   "research_questions", "contributions", "key_concepts", "cited_for"]

    print(f"\n📥 Importing {len(papers)} papers to VF Store...")

    success_count = 0
    error_count = 0

    for idx, paper in enumerate(papers, start=1):
        try:
            profile = paper_to_vf_profile(paper)
            paper_id = profile["paper_id"]
            chunks = profile["chunks"]
            meta = profile["meta"]

            # Prepare texts for embedding
            texts = [chunks.get(chunk_id, "") for chunk_id in CHUNK_ORDER]

            # Encode all chunks
            vectors = model.encode(texts, batch_size=8, show_progress_bar=False)

            # Create points
            points = []
            for chunk_idx, chunk_id in enumerate(CHUNK_ORDER):
                point_id = str(uuid.uuid5(VF_NAMESPACE, f"{paper_id}:{chunk_id}"))
                payload = {
                    "paper_id": paper_id,
                    "chunk_id": chunk_id,
                    "chunk_index": chunk_idx + 1,
                    "in_library": True,
                    "source_type": "slr_excel",
                    "text": texts[chunk_idx],
                    "meta": meta,
                }
                points.append(PointStruct(
                    id=point_id,
                    vector=vectors[chunk_idx].tolist(),
                    payload=payload
                ))

            # Upsert to Qdrant
            client.upsert(collection_name=COLLECTION_NAME, points=points)

            title_short = meta.get("title", "")[:50] + "..." if meta.get("title") else "No title"
            print(f"   [{idx:3d}/{len(papers)}] ✅ {paper_id}: {title_short}")
            success_count += 1

        except Exception as e:
            print(f"   [{idx:3d}/{len(papers)}] ❌ Error: {e}")
            error_count += 1

    return success_count, error_count


def main():
    parser = argparse.ArgumentParser(description="Import SLR Excel data to VF Store")
    parser.add_argument("--start-index", type=int, default=None, help="Start from this index (inclusive)")
    parser.add_argument("--end-index", type=int, default=None, help="End at this index (inclusive)")
    args = parser.parse_args()

    print("=" * 60)
    print("🌸 Veritas SLR Excel Importer")
    print("=" * 60)
    print()

    # Step 1: Read Excel
    papers = read_excel_data(start_index=args.start_index, end_index=args.end_index)

    # Step 2: Setup Qdrant
    client = setup_qdrant()

    # Step 3: Load embedding model
    model = get_embedding_model()

    # Step 4: Import
    success, errors = import_to_vf_store(papers, client, model)

    print()
    print("=" * 60)
    print(f"✨ Import Complete! Range: {args.start_index or 'auto'} - {args.end_index or 'auto'}")
    print(f"   ✅ Success: {success}")
    print(f"   ❌ Errors:  {errors}")
    print(f"   📊 Total vectors: {success * 8} (8 chunks per paper)")
    print(f"   🗄️  Collection: {COLLECTION_NAME}")
    print("=" * 60)


if __name__ == "__main__":
    main()
