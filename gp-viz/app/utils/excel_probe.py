from __future__ import annotations

from dataclasses import dataclass

from openpyxl import load_workbook


@dataclass(frozen=True)
class ExcelInfo:
    exists: bool
    sheet_count: int
    sheet_names: list[str]
    first_sheet_rows: int | None

    def asdict(self) -> dict:
        return {
            "exists": self.exists,
            "sheet_count": self.sheet_count,
            "sheet_names": self.sheet_names,
            "first_sheet_rows": self.first_sheet_rows,
        }


def inspect_excel(path: str) -> ExcelInfo:
    wb = load_workbook(filename=path, read_only=True)
    sheet_names = wb.sheetnames
    first_rows = None
    if sheet_names:
        ws = wb[sheet_names[0]]
        first_rows = ws.max_row
    return ExcelInfo(
        exists=True,
        sheet_count=len(sheet_names),
        sheet_names=sheet_names,
        first_sheet_rows=first_rows,
    )