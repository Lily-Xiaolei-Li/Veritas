from __future__ import annotations

from typing import Any

from openpyxl import load_workbook

from app.utils.config import Settings


def _norm(text: object) -> str:
    return str(text or "").strip().lower()


def _to_int(text: object) -> int | None:
    if text is None:
        return None
    try:
        f = float(str(text).replace(",", "").strip())
        return int(f)
    except (TypeError, ValueError):
        return None


def _to_float(text: object) -> float | None:
    if text is None:
        return None
    try:
        return float(str(text).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _looks_like_header(row: tuple[Any, ...], required: list[str] | None = None) -> int:
    cells = [_norm(c) for c in row if str(c).strip()]
    if len(cells) < 2:
        return 0

    score = len(cells)
    if required:
        req = {_norm(r) for r in required}
        score += sum(1 for c in cells if c in req)

    # prefer rows with a likely x-axis marker such as "year"
    if any("year" in c for c in cells):
        score += 3

    # pivot headers usually include these labels
    if any("row labels" in c for c in cells):
        score += 4
    if any("column labels" in c for c in cells):
        score += 2

    # header rows usually have multiple axis buckets/years
    numeric_like = sum(1 for c in cells if any(ch.isdigit() for ch in c))
    score += min(numeric_like * 2, 8)

    alpha_like = sum(1 for c in cells if any(ch.isalpha() for ch in c))
    score += min(alpha_like, 6)

    return score


def _find_header_row(ws_rows: list[tuple[Any, ...]], required: list[str] | None = None, search_rows: int = 20) -> int:
    best_idx = -1
    best_score = -1
    upper = min(len(ws_rows), search_rows)
    for idx in range(upper):
        score = _looks_like_header(ws_rows[idx], required=required)
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx


def _read_sheet_rows(path: str, sheet_name: str) -> list[tuple[Any, ...]]:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {available}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def load_timeline_data(settings: Settings) -> dict[str, Any]:
    """
    Load timeline chart data from sheet "Stacked area chart".

    The pivot header is usually placed at row 3, so we detect header row by
    heuristics rather than assuming row 1.
    """
    rows = _read_sheet_rows(settings.excel_path, "Stacked area chart")
    header_index = _find_header_row(rows, required=["year", "theme", "category"])
    if header_index < 0:
        raise ValueError("Could not find usable header row in 'Stacked area chart'")

    raw_header = rows[header_index]
    header = [str(raw_header[i]).strip() for i in range(len(raw_header)) if raw_header[i] is not None and str(raw_header[i]).strip()]
    if not header:
        raise ValueError("Header row in 'Stacked area chart' appears empty")

    year_col = next((i for i, v in enumerate(raw_header) if _norm(v) == "year"), 0)
    theme_cols = [(i, str(raw_header[i]).strip()) for i, v in enumerate(raw_header) if i != year_col and v not in (None, "")]

    years: list[int] = []
    values_by_theme: dict[str, list[float]] = {name: [] for _, name in theme_cols}

    for row in rows[header_index + 1 :]:
        y = _to_int(row[year_col] if year_col < len(row) else None)
        if y is None:
            continue
        years.append(y)

        for col_idx, theme in theme_cols:
            value = _to_float(row[col_idx] if col_idx < len(row) else None) or 0.0
            values_by_theme[theme].append(value)

    for theme, vals in values_by_theme.items():
        if len(vals) < len(years):
            vals.extend([0.0] * (len(years) - len(vals)))
        elif len(vals) > len(years):
            values_by_theme[theme] = vals[: len(years)]

    return {
        "sheet": "Stacked area chart",
        "years": years,
        "series": [
            {"name": name, "values": values_by_theme[name]} for _, name in theme_cols
        ],
    }


def _extract_heatmap_headers(rows: list[tuple[Any, ...]]) -> tuple[int, list[str]]:
    """Return (header_row_idx, cleaned_column_headers)."""
    idx = _find_header_row(rows, required=["row labels", "column labels"], search_rows=14)
    if idx < 0:
        return -1, []

    # fallback: pivot title row may be first header row, real header one row down
    raw = [str(v).strip() for v in rows[idx] if v is not None and str(v).strip()]
    if len(raw) < 2 and idx + 1 < len(rows):
        raw_next = [str(v).strip() for v in rows[idx + 1] if v is not None and str(v).strip()]
        if len(raw_next) >= 2:
            idx = idx + 1
            raw = raw_next

    # remove labels like "Column Labels"/"Row Labels" that are pivot artifacts
    if raw:
        if _norm(raw[0]) in {"column labels", "row labels", "label", ""}:
            raw = raw[1:]
        elif len(raw) > 1 and _norm(raw[1]) == "column labels":
            raw = [raw[0]] + raw[2:]

    # keep only non-empty headers after first col
    return idx, raw[1:]


def load_heatmap_sheets(settings: Settings) -> list[dict[str, Any]]:
    """Load heatmap matrices from sheets named like 'Heatmap (theme)'."""
    wb = load_workbook(filename=settings.excel_path, read_only=True, data_only=True)
    target_sheets = [name for name in wb.sheetnames if name.startswith("Heatmap (theme)")]
    result: list[dict[str, Any]] = []

    for sheet_name in target_sheets:
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        if not rows:
            continue

        header_index, col_headers = _extract_heatmap_headers(rows)
        if header_index < 0 or not col_headers:
            continue

        row_labels: list[str] = []
        z: list[list[float]] = []

        for row in rows[header_index + 1 :]:
            if not row:
                continue
            label = str(row[0]).strip() if row[0] is not None else ""
            if not label or _norm(label) in {"row labels", "labels", "column labels"}:
                continue

            values: list[float] = []
            for idx in range(1, len(col_headers) + 1):
                values.append(_to_float(row[idx] if idx < len(row) else None) or 0.0)

            # avoid fully empty trailing rows from pivot export
            if len(values) == 0 or all(v == 0.0 for v in values):
                continue

            if len(values) < len(col_headers):
                values.extend([0.0] * (len(col_headers) - len(values)))
            elif len(values) > len(col_headers):
                values = values[: len(col_headers)]

            row_labels.append(label)
            z.append(values)

        if row_labels and z:
            result.append({"sheet": sheet_name, "x": col_headers, "y": row_labels, "z": z})

    wb.close()
    return result


def load_profile_filter_options(settings: Settings) -> dict[str, Any]:
    """Use vf_profiles_slr (Qdrant) as auxiliary source for filters."""
    try:
        from app.utils.qdrant_client import list_points
    except Exception as exc:  # pragma: no cover
        return {
            "count": 0,
            "availableYears": [],
            "availableJournals": [],
            "availableCountries": [],
            "availablePaperTypes": [],
            "warning": str(exc),
        }

    points = list_points(settings.qdrant_base_url, settings.data_source_collection, limit=4000)

    # deduplicate chunk-level points to paper-level metadata
    dedup: dict[str, dict[str, Any]] = {}
    for point in points:
        payload = point.get("payload") or {}
        meta = payload.get("meta") or {}
        paper_id = str(
            meta.get("index")
            or payload.get("paper_id")
            or payload.get("id")
            or payload.get("source_id")
            or ""
        )
        if not paper_id or paper_id in dedup:
            continue

        dedup[paper_id] = {
            "year": meta.get("year") or payload.get("year"),
            "journal": meta.get("journal") or payload.get("journal") or payload.get("source") or "",
            "country": meta.get("country") or payload.get("country") or "",
            "paper_type": meta.get("paper_type") or payload.get("paper_type") or "",
        }

    available_years = sorted(
        [str(v) for v in (m["year"] for m in dedup.values()) if str(v).strip()],
        key=lambda x: int(x) if x.isdigit() else x,
    )
    available_journals = sorted({str(m["journal"]).strip() for m in dedup.values() if str(m["journal"]).strip()})
    available_countries = sorted({str(m["country"]).strip() for m in dedup.values() if str(m["country"]).strip()})
    available_paper_types = sorted({str(m["paper_type"]).strip() for m in dedup.values() if str(m["paper_type"]).strip()})

    return {
        "count": len(dedup),
        "availableYears": available_years,
        "availableJournals": available_journals,
        "availableCountries": available_countries,
        "availablePaperTypes": available_paper_types,
    }