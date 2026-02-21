"""
Inject canonical_id from Excel index into Library and VF Store.

Logic:
1. Read Excel: filename → canonical_id, item_id mapping
2. For each row, match filename exactly against Library (academic_papers)
3. If matched:
   - Write canonical_id to all Library chunks with that filename
   - If Library chunk has item_id, find VF profile and write canonical_id there too
4. Report stats
"""

import openpyxl
import json
from pathlib import Path
from datetime import datetime
from qdrant_client import QdrantClient
from qdrant_client.models import Filter, FieldCondition, MatchValue, PointStruct

EXCEL_PATH = r"C:\Users\Barry Li (UoN)\OneDrive - The University Of Newcastle\Desktop\AI\Library\Main Library Index (20260117).xlsx"
LIBRARY_COLLECTION = "academic_papers"
VF_COLLECTION = "vf_profiles"
OUTPUT_FILE = Path("data/library_sync/canonical_id_results.json")

def run():
    print("=" * 60)
    print("Inject canonical_id from Excel → Library + VF Store")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Connect
    client = QdrantClient(host="localhost", port=6333)

    # Read Excel
    print("\nReading Excel index...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["filename"]]:
            rows.append({
                "canonical_id": row[idx["canonical_id"]],
                "filename":     row[idx["filename"]],
                "title":        row[idx["title"]],
                "authors":      row[idx["authors"]],
                "year":         row[idx["year"]],
                "doi":          row[idx["doi"]],
            })

    print(f"Excel rows loaded: {len(rows)}")

    # Load ALL Library chunks into memory (filename → point_ids + item_id)
    # NO vectors - use set_payload instead of upsert to avoid memory issues
    print("\nLoading Library (academic_papers)...")
    lib_map = {}  # filename → {"point_ids": [...], "item_id": str|None}
    offset = None
    while True:
        results, offset = client.scroll(
            LIBRARY_COLLECTION,
            limit=500,
            with_payload=True,
            with_vectors=False,  # Don't load vectors - save memory
            offset=offset
        )
        for p in results:
            fn = p.payload.get("filename", "")
            if fn not in lib_map:
                lib_map[fn] = {"point_ids": [], "item_id": p.payload.get("item_id")}
            lib_map[fn]["point_ids"].append(p.id)
        if offset is None:
            break
    print(f"Unique filenames in Library: {len(lib_map)}")

    # Load ALL VF meta chunks (item_id → point_id + paper_id)
    print("Loading VF Store (vf_profiles meta chunks)...")
    vf_map = {}  # item_id → {"point_id": ..., "paper_id": ...}
    offset = None
    while True:
        results, offset = client.scroll(
            VF_COLLECTION,
            scroll_filter=Filter(must=[
                FieldCondition(key="chunk_id", match=MatchValue(value="meta"))
            ]),
            limit=500,
            with_payload=True,
            with_vectors=False,  # Don't load vectors
            offset=offset
        )
        for p in results:
            iid = p.payload.get("meta", {}).get("item_id")
            if iid:
                vf_map[iid] = {"point_id": p.id, "paper_id": p.payload.get("paper_id")}
    print(f"VF profiles with item_id: {len(vf_map)}")

    # Match and inject
    print("\nMatching and injecting canonical_id...")
    print("-" * 60)

    matched_lib = 0
    matched_vf = 0
    not_matched = 0
    results_log = []

    for row in rows:
        excel_fn = row["filename"]
        canonical_id = row["canonical_id"]

        # Try exact filename match in Library
        # Excel filename may or may not have .md extension
        candidates = [excel_fn, excel_fn + ".md"]
        lib_entry = None
        matched_fn = None
        for c in candidates:
            if c in lib_map:
                lib_entry = lib_map[c]
                matched_fn = c
                break

        if not lib_entry:
            not_matched += 1
            results_log.append({
                "canonical_id": canonical_id,
                "excel_filename": excel_fn,
                "status": "not_found_in_library"
            })
            continue

        # Found in Library - inject canonical_id into all chunks via set_payload
        item_id = lib_entry["item_id"]
        point_ids = lib_entry["point_ids"]

        client.set_payload(
            collection_name=LIBRARY_COLLECTION,
            payload={"canonical_id": canonical_id},
            points=point_ids
        )
        matched_lib += 1

        # If this Library entry has item_id, inject canonical_id into VF too
        vf_updated = False
        if item_id and item_id in vf_map:
            vf_point_id = vf_map[item_id]["point_id"]
            client.set_payload(
                collection_name=VF_COLLECTION,
                payload={"canonical_id": canonical_id},
                points=[vf_point_id]
            )
            matched_vf += 1
            vf_updated = True

        print(f"✅ {canonical_id[:40]}")
        print(f"   Library: {matched_fn[:55]}")
        print(f"   item_id: {item_id or 'none'} | VF updated: {vf_updated}")

        results_log.append({
            "canonical_id": canonical_id,
            "excel_filename": excel_fn,
            "library_filename": matched_fn,
            "item_id": item_id,
            "vf_updated": vf_updated,
            "status": "matched"
        })

    # Save log
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(json.dumps(results_log, indent=2, ensure_ascii=False))

    # Summary
    print("\n" + "=" * 60)
    print("COMPLETE")
    print("=" * 60)
    print(f"Excel rows:          {len(rows)}")
    print(f"Matched in Library:  {matched_lib}")
    print(f"Also matched in VF:  {matched_vf}")
    print(f"Not found:           {not_matched}")
    print(f"Match rate:          {100*matched_lib/len(rows):.1f}%")
    print(f"\nResults: {OUTPUT_FILE}")

if __name__ == "__main__":
    run()
