"""Compare Excel index filenames against Library and VF store.
- Excel: Main Library Index (20260117).xlsx
- Library: academic_papers (filename payload)
- VF: vf_profiles meta.source_file

Outputs simple counts, no writes.
"""
import openpyxl
from pathlib import Path
from qdrant_client import QdrantClient
from qdrant_client.models import Filter, FieldCondition, MatchValue

EXCEL_PATH = r"C:\Users\Barry Li (UoN)\OneDrive - The University Of Newcastle\Desktop\AI\Library\Main Library Index (20260117).xlsx"
LIBRARY_COLLECTION = "academic_papers"
VF_COLLECTION = "vf_profiles"

print("=== Excel side ===")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active
headers = [c.value for c in ws[1]]
idx = {h: i for i, h in enumerate(headers)}

excel_fns = []
for row in ws.iter_rows(min_row=2, values_only=True):
    fn = row[idx["filename"]]
    if fn:
        excel_fns.append(fn)

excel_set = set(excel_fns)
print(f"Excel rows: {len(excel_fns)} (unique filenames: {len(excel_set)})")

print("\n=== Library side ===")
client = QdrantClient(host="localhost", port=6333)

lib_fns = set()
offset = None
while True:
    results, offset = client.scroll(
        LIBRARY_COLLECTION,
        limit=500,
        with_payload=True,
        with_vectors=False,
        offset=offset,
    )
    for p in results:
        fn = p.payload.get("filename")
        if fn:
            lib_fns.add(fn)
    if offset is None:
        break

print(f"Library filenames: {len(lib_fns)}")

# exact filename match (with and without .md)
excel_exact_lib = 0
for fn in excel_set:
    if fn in lib_fns or f"{fn}.md" in lib_fns:
        excel_exact_lib += 1

print(f"Excel → Library exact filename match: {excel_exact_lib} / {len(excel_set)}")

print("\n=== VF side (source_file) ===")

vf_sources = set()
offset = None
while True:
    results, offset = client.scroll(
        VF_COLLECTION,
        scroll_filter=Filter(must=[FieldCondition(key="chunk_id", match=MatchValue(value="meta"))]),
        limit=500,
        with_payload=True,
        with_vectors=False,
        offset=offset,
    )
    for p in results:
        src = p.payload.get("meta", {}).get("source_file")
        if src:
            vf_sources.add(src)
    if offset is None:
        break

print(f"VF meta.source_file count: {len(vf_sources)}")

excel_exact_vf = 0
for fn in excel_set:
    if fn in vf_sources or f"{fn}.md" in vf_sources:
        excel_exact_vf += 1

print(f"Excel → VF exact filename match: {excel_exact_vf} / {len(excel_set)}")

print("\n=== Cross check Library & VF ===")
# filenames present in both Library and VF
both = lib_fns & vf_sources
print(f"Filename present in both Library.filename and VF.meta.source_file: {len(both)}")
