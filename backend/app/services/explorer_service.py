"""
Explorer Service - File browsing and import/conversion for Agent B.

Supports:
- Browse any folder on the filesystem
- Import files as Artifacts with automatic conversion
- Excel/CSV → .md (human) + .json (AI) dual format
- Word → .md
- PDF → .md
"""

import os
import json
import logging
from pathlib import Path
from typing import Optional
from datetime import datetime

logger = logging.getLogger(__name__)


# ============================================================
# File Browsing
# ============================================================

def list_directory(path: str) -> dict:
    """
    List contents of a directory.
    Returns folders first, then files, both sorted alphabetically.
    """
    try:
        p = Path(path)
        if not p.exists():
            return {"error": f"Path does not exist: {path}", "items": []}
        if not p.is_dir():
            return {"error": f"Path is not a directory: {path}", "items": []}
        
        items = []
        folders = []
        files = []
        
        for item in p.iterdir():
            try:
                if item.is_dir():
                    try:
                        folder_stat = item.stat()
                        folder_modified = datetime.fromtimestamp(folder_stat.st_mtime).isoformat()
                    except Exception:
                        folder_modified = None
                    folders.append({
                        "name": item.name,
                        "type": "folder",
                        "path": str(item),
                        "modified": folder_modified,
                    })
                else:
                    # Get file info
                    stat = item.stat()
                    files.append({
                        "name": item.name,
                        "type": "file",
                        "path": str(item),
                        "size": stat.st_size,
                        "extension": item.suffix.lower(),
                        "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    })
            except PermissionError:
                continue  # Skip files we can't access
            except Exception as e:
                logger.warning(f"Error reading {item}: {e}")
                continue
        
        # Sort alphabetically, folders first
        folders.sort(key=lambda x: x["name"].lower())
        files.sort(key=lambda x: x["name"].lower())
        items = folders + files
        
        return {
            "path": str(p),
            "parent": str(p.parent) if p.parent != p else None,
            "items": items,
            "count": len(items),
        }
    except Exception as e:
        logger.error(f"Error listing directory {path}: {e}")
        return {"error": str(e), "items": []}


def get_file_info(path: str) -> dict:
    """Get detailed info about a file."""
    try:
        p = Path(path)
        if not p.exists():
            return {"error": f"File does not exist: {path}"}
        if not p.is_file():
            return {"error": f"Path is not a file: {path}"}
        
        stat = p.stat()
        return {
            "name": p.name,
            "path": str(p),
            "size": stat.st_size,
            "extension": p.suffix.lower(),
            "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            "created": datetime.fromtimestamp(stat.st_ctime).isoformat(),
        }
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# File Conversion
# ============================================================

def convert_docx_to_md(file_path: str) -> tuple[str, Optional[str]]:
    """
    Convert Word document to Markdown.
    Returns: (markdown_content, error_message)
    """
    try:
        from docx import Document
        doc = Document(file_path)
        
        lines = []
        for para in doc.paragraphs:
            text = para.text.strip()
            if not text:
                lines.append("")
                continue
            
            # Detect headings by style
            style_name = para.style.name.lower() if para.style else ""
            if "heading 1" in style_name:
                lines.append(f"# {text}")
            elif "heading 2" in style_name:
                lines.append(f"## {text}")
            elif "heading 3" in style_name:
                lines.append(f"### {text}")
            elif "title" in style_name:
                lines.append(f"# {text}")
            else:
                lines.append(text)
        
        return "\n\n".join(lines), None
    except ImportError:
        return "", "python-docx not installed. Run: pip install python-docx"
    except Exception as e:
        return "", f"Error converting Word document: {e}"


def convert_xlsx_to_json(file_path: str) -> tuple[dict, Optional[str]]:
    """
    Convert Excel file to JSON.
    Returns: (json_data, error_message)
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        result = {
            "source_file": Path(file_path).name,
            "sheets": []
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            
            if not rows:
                continue
            
            # First row as headers
            headers = [str(h) if h is not None else f"Column{i}" for i, h in enumerate(rows[0])]
            
            # Data rows
            data_rows = []
            for row in rows[1:]:
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        # Convert to JSON-safe types
                        if val is None:
                            row_dict[headers[i]] = None
                        elif isinstance(val, (int, float)):
                            row_dict[headers[i]] = val
                        else:
                            row_dict[headers[i]] = str(val)
                data_rows.append(row_dict)
            
            result["sheets"].append({
                "name": sheet_name,
                "headers": headers,
                "rows": data_rows,
                "row_count": len(data_rows),
            })
        
        wb.close()
        return result, None
    except ImportError:
        return {}, "openpyxl not installed. Run: pip install openpyxl"
    except Exception as e:
        return {}, f"Error converting Excel file: {e}"


def convert_xlsx_to_md(file_path: str) -> tuple[str, Optional[str]]:
    """
    Convert Excel file to Markdown tables (human-readable).
    Returns: (markdown_content, error_message)
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        md_parts = []
        source_name = Path(file_path).name
        md_parts.append(f"# {source_name}\n")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            
            if not rows:
                continue
            
            md_parts.append(f"\n## {sheet_name}\n")
            
            # First row as headers
            headers = [str(h) if h is not None else "" for h in rows[0]]
            header_row = "| " + " | ".join(headers) + " |"
            separator = "| " + " | ".join(["---"] * len(headers)) + " |"
            
            md_parts.append(header_row)
            md_parts.append(separator)
            
            # Data rows
            for row in rows[1:]:
                cells = [str(v) if v is not None else "" for v in row]
                # Pad to match header count
                while len(cells) < len(headers):
                    cells.append("")
                md_parts.append("| " + " | ".join(cells[:len(headers)]) + " |")
        
        wb.close()
        return "\n".join(md_parts), None
    except ImportError:
        return "", "openpyxl not installed. Run: pip install openpyxl"
    except Exception as e:
        return "", f"Error converting Excel to Markdown: {e}"


def convert_csv_to_json(file_path: str) -> tuple[dict, Optional[str]]:
    """Convert CSV file to JSON."""
    try:
        import csv
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            headers = reader.fieldnames or []
        
        return {
            "source_file": Path(file_path).name,
            "sheets": [{
                "name": "data",
                "headers": headers,
                "rows": rows,
                "row_count": len(rows),
            }]
        }, None
    except Exception as e:
        return {}, f"Error converting CSV: {e}"


def convert_csv_to_md(file_path: str) -> tuple[str, Optional[str]]:
    """Convert CSV file to Markdown table."""
    try:
        import csv
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            rows = list(reader)
        
        if not rows:
            return "", "CSV file is empty"
        
        source_name = Path(file_path).name
        md_parts = [f"# {source_name}\n"]
        
        headers = rows[0]
        header_row = "| " + " | ".join(headers) + " |"
        separator = "| " + " | ".join(["---"] * len(headers)) + " |"
        
        md_parts.append(header_row)
        md_parts.append(separator)
        
        for row in rows[1:]:
            cells = [str(v) for v in row]
            while len(cells) < len(headers):
                cells.append("")
            md_parts.append("| " + " | ".join(cells[:len(headers)]) + " |")
        
        return "\n".join(md_parts), None
    except Exception as e:
        return "", f"Error converting CSV to Markdown: {e}"


def convert_pdf_to_md(file_path: str) -> tuple[str, Optional[str]]:
    """Convert PDF to Markdown (cleaned text extraction).

    Stage 6 improvements:
    - remove repeated headers/footers
    - fix hard line breaks → paragraphs
    - de-hyphenate words split across lines
    - promote obvious headings

    Returns: (markdown_content, error_message)
    """

    try:
        import re
        from collections import Counter

        import fitz  # PyMuPDF

        def _normalize_line(s: str) -> str:
            # normalize digits/spaces so "Page 3" and "Page 4" match
            s = re.sub(r"\d+", "#", s)
            s = re.sub(r"\s+", " ", s).strip()
            return s

        def _looks_like_page_number(s: str) -> bool:
            s2 = s.strip()
            if not s2:
                return False
            return bool(re.fullmatch(r"(page\s+)?\d+(\s+of\s+\d+)?", s2, flags=re.I))

        def _is_heading(line: str) -> bool:
            s = line.strip()
            if not s:
                return False
            if len(s) > 90:
                return False
            # numbered headings: 1. / 1.2 / 2.3.4
            if re.match(r"^\d+(?:\.\d+)*\s+\S+", s):
                return True
            # ALL CAPS headings (with some letters)
            letters = re.sub(r"[^A-Za-z]", "", s)
            if letters and letters.isupper() and len(letters) >= 6:
                return True
            return False

        def _heading_level(line: str) -> int:
            s = line.strip()
            m = re.match(r"^(\d+(?:\.\d+)*)\s+", s)
            if m:
                depth = m.group(1).count(".") + 1
                return 2 if depth <= 1 else 3
            return 2

        def _join_lines_to_paragraphs(lines: list[str]) -> list[str]:
            out: list[str] = []
            buf: list[str] = []

            def flush():
                if buf:
                    out.append(" ".join(buf).strip())
                    buf.clear()

            i = 0
            while i < len(lines):
                line = lines[i].rstrip()
                s = line.strip()

                if not s:
                    flush()
                    i += 1
                    continue

                if _looks_like_page_number(s):
                    i += 1
                    continue

                if _is_heading(s):
                    flush()
                    level = _heading_level(s)
                    out.append(f"{'#' * level} {s}")
                    i += 1
                    continue

                # de-hyphenate across line break
                if s.endswith("-") and i + 1 < len(lines):
                    nxt = lines[i + 1].lstrip()
                    if nxt and nxt[:1].islower():
                        s = s[:-1] + nxt
                        i += 2
                        # continue processing merged string as part of paragraph
                        buf.append(s)
                        continue

                # paragraph joining heuristic
                if buf:
                    prev = buf[-1]
                    if prev and not prev.endswith(('.', '!', '?', ':', ';')) and s[:1].islower():
                        buf.append(s)
                    else:
                        # likely new sentence/paragraph
                        flush()
                        buf.append(s)
                else:
                    buf.append(s)

                i += 1

            flush()
            return out

        doc = fitz.open(file_path)
        source_name = Path(file_path).name

        # collect per-page lines first
        pages: list[list[str]] = []
        header_footer_counter: Counter[str] = Counter()
        header_footer_candidates: list[str] = []

        for page in doc:
            text = page.get_text("text") or ""
            raw_lines = [ln.rstrip() for ln in text.splitlines()]
            pages.append(raw_lines)

            # gather top/bottom lines as header/footer candidates
            trimmed = [ln.strip() for ln in raw_lines if ln.strip()]
            top = trimmed[:3]
            bottom = trimmed[-3:] if len(trimmed) >= 3 else trimmed
            for ln in top + bottom:
                n = _normalize_line(ln)
                if n:
                    header_footer_counter[n] += 1

        page_count = max(1, len(pages))
        # consider repeated if appears on >=60% of pages
        threshold = max(2, int(page_count * 0.6))
        repeated = {k for k, v in header_footer_counter.items() if v >= threshold}

        md_parts: list[str] = [f"# {source_name}\n"]

        for idx, raw_lines in enumerate(pages, 1):
            cleaned_lines: list[str] = []
            for ln in raw_lines:
                s = ln.strip()
                if not s:
                    cleaned_lines.append("")
                    continue
                if _normalize_line(s) in repeated:
                    continue
                cleaned_lines.append(ln)

            blocks = _join_lines_to_paragraphs(cleaned_lines)
            if not blocks:
                continue

            # keep a lightweight page marker (less noisy than big headings)
            md_parts.append(f"\n<!-- Page {idx} -->\n")
            md_parts.extend(blocks)

        doc.close()
        return "\n\n".join(md_parts).strip() + "\n", None

    except ImportError:
        return "", "PyMuPDF not installed. Run: pip install PyMuPDF"
    except Exception as e:
        return "", f"Error converting PDF: {e}"


def read_text_file(file_path: str) -> tuple[str, Optional[str]]:
    """Read a text file (txt, md, json, etc)."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read(), None
    except UnicodeDecodeError:
        try:
            with open(file_path, 'r', encoding='latin-1') as f:
                return f.read(), None
        except Exception as e:
            return "", f"Error reading file: {e}"
    except Exception as e:
        return "", f"Error reading file: {e}"


# ============================================================
# Import Logic (Payload generation)
# ============================================================

def convert_file_to_artifact_payloads(file_path: str) -> list[dict]:
    """Convert a file into one or more artifact payloads.

    Backward-compatible wrapper: conversion logic lives in
    `document_processing_service` so other ingestion paths can reuse it.
    """

    from .document_processing_service import convert_file_to_artifact_payloads as _convert

    return _convert(file_path, imported_via="explorer")
